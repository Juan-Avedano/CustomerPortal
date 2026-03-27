# admin/views.py
from flask import (Blueprint, render_template, flash, redirect, url_for,request, jsonify, current_app, send_file)
from flask_login import login_required, current_user
from models import db, User, Service, Complaint, user_services_association, TechnicalVisit, Technician, Peticion, CalificacionServicio, Invoice
from forms import UserEditForm, ToggleUserStatusForm, AssignTechnicianForm

from werkzeug.security import generate_password_hash
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from decorators import admin_required
from datetime import datetime
from utils import create_notification 
from sqlalchemy import func, desc
import pandas as pd
import io


admin_bp = Blueprint('admin', __name__, 
                     template_folder='../templates/admin', 
                     url_prefix='/admin')

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    return render_template('admin_dashboard.html', active_tab='overview')

#Ruta exportar excel
@admin_bp.route('/user/<int:user_id>/exportar-excel')
@login_required
@admin_required
def exportar_excel_usuario(user_id):
    user = User.query.get_or_404(user_id)
    user_services = user.services_contracted.all()

    if not user_services:
        flash(f"El usuario '{user.username}' no tiene servicios para exportar.", 'warning')
        return redirect(url_for('admin.user_detail', user_id=user_id))

    datos_servicios = []
    total_amount = 0.0
    for service in user_services:
        datos_servicios.append({
            'ID Servicio': service.id, 'Nombre Servicio': service.name,
            'Tipo': service.type, 'Precio': service.price
        })
        total_amount += service.price
    df = pd.DataFrame(datos_servicios)

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    df_cliente = pd.DataFrame([
        {'Campo': 'ID Cliente', 'Valor': user.id},
        {'Campo': 'Nombre de Usuario', 'Valor': user.username},
        {'Campo': 'Email', 'Valor': user.email},
        {'Campo': 'Fecha de Exportación', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')}
    ])
    df_cliente.to_excel(writer, sheet_name='Info Cliente', index=False, startrow=1)
    df.to_excel(writer, sheet_name='Servicios Contratados', index=False, startrow=1)

    workbook = writer.book
    ws_cliente = workbook['Info Cliente']
    ws_servicios = workbook['Servicios Contratados']
    
    # --- Estilos de Excel ---
    font_titulo = Font(name='Calibri', size=16, bold=True)
    font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    alineacion_centrada = Alignment(horizontal='center', vertical='center')
    relleno_header = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    formato_moneda = '$ #,##0.00'
    ws_servicios.insert_rows(1); ws_servicios.merge_cells('A1:D1'); titulo_servicios = ws_servicios['A1']
    titulo_servicios.value = f"Reporte de Servicios de: {user.username}"; titulo_servicios.font = font_titulo; titulo_servicios.alignment = alineacion_centrada
    for cell in ws_servicios[2]: cell.font = font_header; cell.fill = relleno_header; cell.border = borde_fino; cell.alignment = alineacion_centrada
    for row in ws_servicios.iter_rows(min_row=3, max_row=ws_servicios.max_row, min_col=1, max_col=4):
        for cell in row: cell.border = borde_fino
        row[3].number_format = formato_moneda; row[3].alignment = Alignment(horizontal='right', vertical='center')
    total_row_index = ws_servicios.max_row + 1; ws_servicios[f'C{total_row_index}'] = 'TOTAL:'; ws_servicios[f'D{total_row_index}'] = total_amount
    total_label_cell = ws_servicios[f'C{total_row_index}']; total_value_cell = ws_servicios[f'D{total_row_index}']
    total_label_cell.font = Font(name='Calibri', size=11, bold=True); total_label_cell.alignment = Alignment(horizontal='right', vertical='center'); total_label_cell.border = borde_fino
    total_value_cell.font = Font(name='Calibri', size=11, bold=True); total_value_cell.number_format = formato_moneda; total_value_cell.border = borde_fino
    for i, column_cells in enumerate(ws_servicios.columns, 1):
        column_letter = get_column_letter(i);
        if i == 4: ws_servicios.column_dimensions[column_letter].width = 15; continue
        max_length = 0
        for cell in column_cells:
            if isinstance(cell, MergedCell): continue
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws_servicios.column_dimensions[column_letter].width = (max_length + 2)
    ws_cliente.insert_rows(1); ws_cliente.merge_cells('A1:B1'); titulo_cliente = ws_cliente['A1']
    titulo_cliente.value = "Información del Cliente"; titulo_cliente.font = font_titulo; titulo_cliente.alignment = alineacion_centrada
    for cell in ws_cliente[2]: cell.font = font_header; cell.fill = relleno_header; cell.border = borde_fino
    for row in ws_cliente.iter_rows(min_row=3, max_row=ws_cliente.max_row, min_col=1, max_col=2):
        for cell in row: cell.border = borde_fino
    ws_cliente.column_dimensions['A'].width = 25; ws_cliente.column_dimensions['B'].width = 40
    
    writer.close()
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'Reporte_{user.username}.xlsx')

@admin_bp.route('/users')
@login_required
@admin_required
def manage_users():
    users = User.query.all()
    return render_template('admin_user.html', users=users, active_tab='users')

@admin_bp.route('/user/<int:user_id>')
@login_required
@admin_required
def user_detail(user_id):
    user = User.query.get_or_404(user_id)
    toggle_form=ToggleUserStatusForm()
    return render_template('admin_detail.html', user=user, user_services=user.services_contracted.all(), active_tab='users', toggle_form=toggle_form)

@admin_bp.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = UserEditForm(obj=user)
    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        flash(f'✅ El usuario "{user.username}" ha sido actualizado exitosamente.', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    return render_template('admin_edit_user.html', form=form, user=user, active_tab='users')

@admin_bp.route('/user/<int:user_id>/toggle_status', methods=['POST'])
@login_required
@admin_required
def toggle_user_active_status(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin or user.id == current_user.id:
        flash('No se puede cambiar el estado de este usuario.', 'danger')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        flash(f'Estado de {user.username} cambiado.', 'success')
    return redirect(url_for('admin.user_detail', user_id=user.id))

@admin_bp.route('/user/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin or user.id == current_user.id:
        flash('No se puede eliminar a este usuario.', 'danger')
        return redirect(url_for('admin.manage_users'))
    db.session.delete(user)
    db.session.commit()
    flash('Usuario eliminado.', 'success')
    return redirect(url_for('admin.manage_users'))




# === RUTAS PARA LA GESTIÓN DE VISITAS TÉCNICAS ===

@admin_bp.route('/visits')
@login_required
@admin_required
def manage_visits():
    pending_visits = TechnicalVisit.query.filter_by(status='Agendada').order_by(TechnicalVisit.scheduled_datetime.asc()).all()
    return render_template('admin_manage_visits.html', visits=pending_visits, active_tab='visits')

@admin_bp.route('/visit/<int:visit_id>/assign', methods=['GET', 'POST'])
@login_required
@admin_required
def assign_technician(visit_id):
    visit = TechnicalVisit.query.get_or_404(visit_id)
    form = AssignTechnicianForm()
    technicians = Technician.query.all()
    form.technician.choices = [(t.id, f"{t.name} ({t.specialty or 'General'})") for t in technicians]
    
    if form.validate_on_submit():
        technician_id = form.technician.data
        visit.technician_id = technician_id
        visit.status = 'En Camino'
        
        create_notification(
            visit.user,
            f"Se ha asignado un técnico a tu visita para '{visit.service.name}'.",
            link=url_for('visits.list_visits')
        )
        
        db.session.commit()
        flash(f"Técnico asignado a la visita #{visit.id} exitosamente.", 'success')
        return redirect(url_for('admin.manage_visits'))

    return render_template('admin_assign_technician.html', visit=visit, form=form, active_tab='visits')


@admin_bp.route('/visit/<int:visit_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_visit(visit_id):
    """
    Permite a un administrador eliminar una visita técnica permanentemente.
    """
    visit_to_delete = TechnicalVisit.query.get_or_404(visit_id)
    
    # Guardamos el ID para el mensaje flash antes de eliminar
    visit_id_for_flash = visit_to_delete.id
    
    db.session.delete(visit_to_delete)
    db.session.commit()
    
    flash(f'La visita #{visit_id_for_flash} ha sido eliminada permanentemente.', 'success')
    return redirect(url_for('admin.manage_visits'))
# --- RUTAS DE API ---

@admin_bp.route('/api/feedback_trend')
@login_required
@admin_required
def feedback_trend_api():
    """Devuelve las últimas calificaciones y fechas para un gráfico de tendencia."""
    
    # 1. Consulta: Obtener las últimas 10 calificaciones ordenadas por fecha
    ratings = CalificacionServicio.query.order_by(
        desc(CalificacionServicio.fecha_calificacion)
    ).limit(10).all()
    
    # Invertir el orden para que la tendencia vaya de izquierda (viejo) a derecha (nuevo)
    ratings.reverse() 

    # 2. Formatear la data
    # Etiquetas: Días/Mes de la calificación
    labels = [r.fecha_calificacion.strftime('%d/%b') for r in ratings]
    # Datos: El valor de la puntuación (0-10)
    data = [r.puntuacion for r in ratings]
    
    return jsonify({
        'labels': labels,
        'data': data
    })






@admin_bp.route('/api/visits_status_distribution')
@login_required
@admin_required
def visits_status_distribution_api():
    """Devuelve el conteo de Visitas Técnicas agrupadas por su estado."""
    
    # Consulta: Contar visitas y agruparlas por su estado
    visit_stats = db.session.query(
        TechnicalVisit.status, 
        func.count(TechnicalVisit.id).label('count')
    ).group_by(TechnicalVisit.status).all()
    
    # Formatear la data
    labels = [stat.status for stat in visit_stats]
    data = [stat.count for stat in visit_stats]
    
    # Definición de colores para cada estado (clave para la consistencia)
    color_map = {
        'Agendada': '#ffc107',      # Amarillo/Naranja (Pendiente)
        'En Camino': '#0d6efd',     # Azul (Activo/En progreso)
        'Completada': '#198754',    # Verde (Éxito)
        'Cancelada': '#dc3545',     # Rojo (Fallo/Cancelación)
        'Asignada': '#6f42c1',      # Púrpura
    }

    return jsonify({
        'labels': labels,
        'data': data,
        'background_colors': [color_map.get(label, '#6c757d') for label in labels]
    })




@admin_bp.route('/api/pack_requests_status')
@login_required
@admin_required
def pack_requests_status_api():
    """Devuelve el conteo de Peticiones de packs (fútbol, premium, etc.) por estado."""
    
    # Consulta: Contar peticiones y agruparlas por su estado
    request_stats = db.session.query(
        Peticion.estado, 
        func.count(Peticion.id).label('count')
    ).group_by(Peticion.estado).all()
    
    # Formatear la data
    labels = [stat.estado for stat in request_stats]
    data = [stat.count for stat in request_stats]
    
    # Definir colores según el estado de la gestión
    color_map = {
        'Aplicado': '#198754',      # Verde (Éxito de Venta/Activación)
        'Pendiente': '#ffc107',     # Amarillo (A la espera de gestión)
        'Rechazado': '#dc3545',     # Rojo (Venta fallida/No procede)
    }

    return jsonify({
        'labels': labels,
        'data': data,
        'background_colors': [color_map.get(label, '#6c757d') for label in labels]
    })



@admin_bp.route('/api/user_stats')
@login_required
@admin_required
def user_stats_api():
    active_users = User.query.filter_by(is_active=True).count()
    inactive_users = User.query.filter_by(is_active=False).count()
    return jsonify({'labels': ['Usuarios Activos', 'Usuarios Inactivos'], 'data': [active_users, inactive_users]})

@admin_bp.route('/api/service_stats')
@login_required
@admin_required
def service_stats_api():
    service_data = db.session.query(
        Service.name, func.count(user_services_association.c.user_id)
    ).join(user_services_association).group_by(Service.name).order_by(func.count(user_services_association.c.user_id).desc()).all()
    
    return jsonify({
        'labels': [row[0] for row in service_data],
        'data': [row[1] for row in service_data]
    })
    


@admin_bp.route('/peticiones')
@login_required
@admin_required
def listar_peticiones():
    """
    Permite al administrador consultar el listado de peticiones de planes enviadas por los clientes.
    """
    # Filtro por defecto: 'Pendiente'
    estado_filtro = request.args.get('estado', 'Pendiente')
    
    query = Peticion.query.order_by(desc(Peticion.fecha_creacion))
    
    # Aplica el filtro de estado si no es 'Todas'
    if estado_filtro != 'Todas':
        query = query.filter(Peticion.estado == estado_filtro)
    
    peticiones = query.all()
    
    # Contadores para mostrar en la interfaz
    contadores = {
        'Pendiente': Peticion.query.filter_by(estado='Pendiente').count(),
        'Aplicado': Peticion.query.filter_by(estado='Aplicado').count(),
        'Eliminada': Peticion.query.filter_by(estado='Eliminada').count(),
        'Rechazado': Peticion.query.filter_by(estado='Rechazado').count(),
        'Total': Peticion.query.count()
    }
    
    return render_template('admin_manage_peticiones.html', 
                           title='Gestión de Peticiones', 
                           peticiones=peticiones, 
                           estado_filtro=estado_filtro,
                           contadores=contadores,
                           active_tab='peticiones')


@admin_bp.route('/peticiones/gestionar/<int:peticion_id>', methods=['POST'])
@login_required
@admin_required
def gestionar_peticion(peticion_id):
    """
    Permite al administrador modificar el estado de una petición.
    (RD: Modificar/Eliminar peticiones enviadas)
    """
    peticion = Peticion.query.get_or_404(peticion_id)
    
    # La acción viene del formulario (puede ser 'aplicar', 'eliminar' o 'rechazar')
    nueva_accion = request.form.get('accion') 
    
    if not nueva_accion:
        flash('Acción no especificada.', 'danger')
        return redirect(url_for('admin.listar_peticiones'))

    try:
        peticion.fecha_gestion = datetime.utcnow()
        
        if nueva_accion == 'aplicar':
            peticion.estado = 'Aplicado'
            flash(f"Petición #{peticion.id} ({peticion.detalle_plan}) APLICADA.", 'success')
            create_notification(peticion.user, f"Su solicitud para '{peticion.detalle_plan}' ha sido APLICADA.", link=url_for('customer_dashboard.ver_peticiones'))
            
        elif nueva_accion == 'eliminar':
            peticion.estado = 'Eliminada'
            flash(f"Petición #{peticion.id} marcada como ELIMINADA.", 'warning')
            
        elif nueva_accion == 'rechazar':
            peticion.estado = 'Rechazado'
            flash(f"Petición #{peticion.id} marcada como RECHAZADA.", 'warning')

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al gestionar la petición: {e}', 'danger')
        
    # Redirige al filtro que estaba activo o al de Pendientes
    
    return redirect(url_for('admin.listar_peticiones', estado=peticion.estado))


@admin_bp.route('/feedback/calificaciones')
@login_required
@admin_required
def ver_calificaciones():
    """
    Permite al administrador consultar y analizar el feedback de las visitas técnicas.
    """
    
    # 1. Análisis de Métricas Clave
    total_visitas = CalificacionServicio.query.count()
    
    # Cálculo del promedio general de puntuación
    promedio_puntuacion = db.session.query(func.avg(CalificacionServicio.puntuacion)).scalar()
    
    # Cálculo del porcentaje de satisfacción válida (problema resuelto)
    resueltas = CalificacionServicio.query.filter_by(satisfaccion_valida=True).count()
    porcentaje_resuelto = (resueltas / total_visitas) * 100 if total_visitas > 0 else 0
    
    analisis = {
        'total': total_visitas,
        'promedio': round(promedio_puntuacion, 2) if promedio_puntuacion else 0,
        'resueltas': round(porcentaje_resuelto, 1)
    }

    # 2. Listado Detallado: Unimos las tablas para obtener el nombre del cliente y del técnico
    calificaciones_con_detalle = db.session.query(
        CalificacionServicio, 
        TechnicalVisit, 
        User # El usuario que calificó
    ).join(TechnicalVisit, CalificacionServicio.technical_visit_id == TechnicalVisit.id) \
     .join(User, CalificacionServicio.user_id == User.id) \
     .order_by(desc(CalificacionServicio.fecha_calificacion)).all()

    return render_template('admin_feedback_report.html', 
                           title='Reporte de Feedback Técnico', 
                           calificaciones=calificaciones_con_detalle,
                           analisis=analisis,
                           active_tab='feedback')
    
    
@admin_bp.route('/visits/audit')
@login_required
@admin_required
def auditar_visitas():
    """
    Permite al administrador visualizar todas las visitas técnicas históricas 
    y filtrarlas por estado (Agendadas, Completadas, Canceladas).
    """
    # Filtro para ver el estado de la visita (por defecto, muestra las más recientes)
    estado_filtro = request.args.get('estado', 'Todas')
    
    query = TechnicalVisit.query.order_by(desc(TechnicalVisit.scheduled_datetime))
    
    if estado_filtro != 'Todas':
        query = query.filter(TechnicalVisit.status == estado_filtro)
    
    visitas = query.all()
    
    # Contadores para la interfaz de auditoría
    contadores = {
        'Agendada': TechnicalVisit.query.filter_by(status='Agendada').count(),
        'En Camino': TechnicalVisit.query.filter_by(status='En Camino').count(),
        'Completada': TechnicalVisit.query.filter_by(status='Completada').count(),
        'Cancelada': TechnicalVisit.query.filter_by(status='Cancelada').count(),
        'Total': TechnicalVisit.query.count()
    }
    
    return render_template('admin_audit_visits.html', 
                           title='Auditoría de Visitas Técnicas', 
                           visitas=visitas, 
                           estado_filtro=estado_filtro,
                           contadores=contadores,
                           active_tab='auditoria')
    
    
    
    
    

# --- Reporte 1: Adopción del Portal (Estratégico) ---
@admin_bp.route('/api/reportes/adopcion')
@login_required
@admin_required
def reporte_adopcion():
    # SIMULACIÓN de datos para el gráfico de línea
    labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun"]
    

    
    gestiones_autogestionadas = [450, 480, 520, 600, 650, 700] 
    gestiones_asistidas = [150, 145, 130, 110, 105, 90]

    return jsonify({
        'labels': labels,
        'autogestionadas': gestiones_autogestionadas,
        'asistidas': gestiones_asistidas
    })

# --- Reporte 2: Eficiencia de Soporte (Táctico) ---
@admin_bp.route('/api/reportes/eficiencia_soporte')
@login_required
@admin_required
def reporte_eficiencia_soporte():
    # SIMULACIÓN de datos para el gráfico de Donut
    
    # ⚠️ NOTA: En un sistema real, esto se obtendría del log del chatbot.
    resueltas_bot = 850 
    derivadas_whatsapp = 150
    
    return jsonify({
        'labels': ['Resuelto por ChatBot', 'Derivado a WhatsApp'],
        'data': [resueltas_bot, derivadas_whatsapp]
    })

# --- Reporte 3: Gestión de Pagos (Estratégico/Financiero) ---
@admin_bp.route('/api/reportes/gestion_pagos')
@login_required
@admin_required
def reporte_gestion_pagos():
    # SIMULACIÓN DE DATOS REALES (Debería usar la tabla Invoice)
    labels = ["Q1 2025", "Q2 2025", "Q3 2025"] # Trimestres
    
    # Monto de facturas pagadas (status='Pagada', external_reference IS NOT NULL)
    recaudado = [950000, 1100000, 1300000]
    
    # Monto de facturas pendientes/vencidas (status='Pendiente' y fecha antigua)
    deuda_vencida = [150000, 120000, 180000]

    return jsonify({
        'labels': labels,
        'recaudado': recaudado,
        'deuda_vencida': deuda_vencida
    })

# --- Reporte 4: Calidad de Servicio (Estratégico) ---
@admin_bp.route('/api/reportes/calidad_servicio')
@login_required
@admin_required
def reporte_calidad_servicio():
    # Usa CalificacionServicio y agrupa por el tipo de servicio
    
    # SIMULACIÓN de resultados de CSAT/NPS por tipo de servicio
    labels = ["Internet Fibra", "TV Premium", "Telefonía Fija"]
    puntuacion_promedio = [8.5, 9.2, 7.8]

    return jsonify({
        'labels': labels,
        'puntuacion': puntuacion_promedio
    })